# encoding:utf-8
import os
import sys
import glob
import time
import openpyxl
import pandas as pd

# 合并后的文件名称前缀
MERGED_FILE_NAME_PREFIX = 'merge_result_'
# excel文件所在的目录
base_path = ''


def merge_file_path():
    """
    :return: 获取合并后的文件绝对路径, 输出目录不存在会进行创建
    """
    output_dir = os.path.join(base_path, 'output')
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)

    return os.path.join(output_dir, MERGED_FILE_NAME_PREFIX +
                        time.strftime("%Y%m%d%H%M%S", time.localtime()) + '.xlsx')


def transformed_file_name(original_name):
    """
    获取转换之后的文件名
    :param original_name: 原始文件名, 不带后缀
    :return:  转换为xlsx之后的文件名
    """
    return original_name + "_" + MERGED_FILE_NAME_PREFIX + time.strftime("%Y%m%d%H%M%S",
                                                                         time.localtime()) + "_trans.xlsx"


def merge_xlsx_files(xlsx_file_list):
    """
    合并传入的xlsx文件, 第一行当做表头处理, 输出到 merge_result_...xlsx 文件
    :param xlsx_file_list:
    :return: None
    """
    xlsx_count = len(xlsx_file_list)
    if xlsx_count < 2:
        print('files count less than two, not need merged')
        return

    # 读取第一个Excel, 使用它的表头
    print('load first file %s' % xlsx_file_list[0])
    merged_workbook = openpyxl.load_workbook(xlsx_file_list[0])
    # 获取活跃的Sheet
    merged_sheet = merged_workbook.active

    # 汇总的Sheet名称
    merged_sheet.title = "merge_result"

    # 遍历除第一个之外的Excel文件
    for filename in xlsx_file_list[1:]:
        print("append file %s start" % filename)
        workbook = openpyxl.load_workbook(filename, data_only=True)
        # 获取每个Excel文件中活跃的工作表
        sheet = workbook.active

        # 从第二行开始读取, 第一行当做表头处理
        for row in sheet.iter_rows(min_row=2):
            # 获取所有单元格的值
            values = [cell.value for cell in row]
            # 空单元格的数量
            none_cell_count = values.count(None)
            # 所有单元格数量
            total_cell_count = len(values)
            # 非空单元格数量
            non_null_count = total_cell_count - none_cell_count
            # 如果全都是空单元格, 跳过
            if non_null_count < 1:
                continue
            merged_sheet.append(values)
        print("append file {} finish".format(filename))

    # 输出文件
    merged_file = merge_file_path()
    print("save file to disk %s" % merged_file)
    merged_workbook.save(merged_file)


def list_all_xlsx_files():
    """
    获取 base_path 下所有 xlsx 文件, 并按照文件名称排序
    :return:
    """
    xlsx_file_list = glob.glob(os.path.join(base_path, '*.xlsx'))
    sorted(xlsx_file_list, key=str.lower)
    return xlsx_file_list


def trans_format():
    """
    转换 base_path 下所有 xls 文件为 xlsx 文件, 保存到一个临时目录下
    :return: None
    """
    # 获取给定目录下的所有文件
    files_list = os.listdir(base_path)
    for file in files_list:
        # 对路径进行分割，分别为文件路径和文件后缀
        file_name_original, suffix = os.path.splitext(file)
        if suffix == '.xls':
            print('now trans format {}...'.format(file))
            # 读取xls文件
            data = pd.DataFrame(pd.read_excel(file))
            # 格式转换
            data.to_excel(os.path.join(base_path, 'trans_temp', file_name_original + "_trans.xlsx"),
                          index=False)
            print(' {} transformed {} saved at {}\n'.format(file, file_name_original + '_trans.xlsx', base_path))


def set_base_path():
    """
    根据参数给base_path赋值, 如果没有默认使用当前目录
    :return:
    """
    global base_path
    if len(sys.argv) > 1:
        base_path = sys.argv[1]
    else:
        base_path = os.getcwd()
    print('base_path %s' % base_path)


if __name__ == '__main__':
    set_base_path()
    # 获取所有 xlsx 文件
    xlsx_files = list_all_xlsx_files()
    # 合并文件
    merge_xlsx_files(xlsx_files)
    # python main.py "C:\Users\ShiLY\Desktop"
